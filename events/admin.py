from django.contrib import admin
from django.utils.html import format_html
from django.utils import timezone
from django.utils.formats import date_format
from .models import Event
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from django.http import HttpResponse
from io import BytesIO
from collections import defaultdict


def _fmt_dt(dt):
    if not dt:
        return "—"
    dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return date_format(dt, "j E Y г. H:i")


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ("card",)
    list_display_links = ("card",)
    list_per_page = 20

    search_fields = (
        "name",
        "comment",
        "responsible",
        "place",
        "category__name",
        "department__name",
    )

    list_filter = (
        "category",
        "department",
        "date",
    )

    class Media:
        css = {"all": ("events/admin-card.css",)}

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("category", "department", "user").distinct()

    def card(self, obj: Event):
        return format_html(
            """
            <div class="mc-card">
              <div class="mc-head">
                <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
                  <span class="mc-id">#{id}</span>
                  <h2 class="mc-title">{title}</h2>
                </div>
              </div>

              <div class="mc-grid">
                <div class="mc-box"><div class="mc-label">Даты</div><p class="mc-val">{date} — {end_date}</p></div>
                <div class="mc-box"><div class="mc-label">Категория</div><p class="mc-val">{category}</p></div>
                <div class="mc-box"><div class="mc-label">Подразделение</div><p class="mc-val">{department}</p></div>
                <div class="mc-box"><div class="mc-label">Ответственный</div><p class="mc-val">{responsible}</p></div>
                <div class="mc-box"><div class="mc-label">Место</div><p class="mc-val">{place}</p></div>
                <div class="mc-box"><div class="mc-label">Создатель</div><p class="mc-val">{creator}</p></div>
              </div>

              <div class="mc-box" style="margin-top:12px">
                <div class="mc-label">Комментарий</div>
                <div class="mc-prose">{comment}</div>
              </div>
            </div>
            """,
            id=obj.pk,
            title=obj.name or "—",
            date=_fmt_dt(obj.date),
            end_date=_fmt_dt(obj.end_date),
            category=getattr(obj.category, "name", "—"),
            department=getattr(obj.department, "name", "—"),
            responsible=obj.responsible or "—",
            place=obj.place or "—",
            creator = obj.user.displayname if obj.user else "—",
            comment=(obj.comment or "—").replace("\n", "<br/>"),
        )

    card.short_description = "Мероприятие"

    # Экспорт в Excel
    def export_to_excel(self, request, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "План мероприятий"

        column_widths = {
            'A': 15,
            'C': 40,
            'D': 35,
            'E': 30,
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        for _ in range(3):
            current_row += 1

        worksheet.merge_cells(f'D{current_row}:E{current_row}')
        worksheet.cell(row=current_row, column=4, value='УТВЕРЖДАЮ').alignment = center_alignment
        worksheet.cell(row=current_row, column=4).font = Font(bold=True)
        current_row += 1

        worksheet.merge_cells(f'D{current_row}:E{current_row}')
        worksheet.cell(row=current_row, column=4, value='Ректор ВоГУ').alignment = center_alignment
        current_row += 1

        worksheet.merge_cells(f'D{current_row}:E{current_row}')
        worksheet.cell(row=current_row, column=4, value='').alignment = center_alignment
        current_row += 1

        worksheet.cell(row=current_row, column=4, value='').alignment = center_alignment
        worksheet.cell(row=current_row, column=5, value='Д. В. Дворников').alignment = center_alignment
        current_row += 1

        worksheet.cell(row=current_row, column=4, value='').alignment = center_alignment
        worksheet.cell(row=current_row, column=5, value='2025 года').alignment = center_alignment
        current_row += 2

        worksheet.merge_cells(f'A{current_row}:E{current_row}')
        worksheet.cell(row=current_row, column=1, value='ПЛАН МЕРОПРИЯТИЙ УНИВЕРСИТЕТА')
        worksheet.cell(row=current_row, column=1).alignment = center_alignment
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2

        headers = [
            "Дата, время",
            "Наименование мероприятия",
            "Место проведения мероприятия",
            "Структурное подразделение",
            "Ответственный работник"
        ]

        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            # Серая заливка для заголовков
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        current_row += 1

        events_by_category = defaultdict(list)

        for event in queryset.order_by('category__name', 'date'):
            if event.category:
                events_by_category[event.category.name].append(event)
            else:
                events_by_category['Без категории'].append(event)

        for category_name, events in events_by_category.items():
            # Добавляем строку с названием категории
            if category_name and events:
                worksheet.merge_cells(f'A{current_row}:E{current_row}')
                category_cell = worksheet.cell(row=current_row, column=1, value=category_name.upper())
                category_cell.font = Font(bold=True, size=11)
                category_cell.alignment = left_alignment
                current_row += 1

                for event in events:
                    date_str = ""
                    if event.date:
                        date_str = event.date.strftime('%Y.%m.%d %H:%M')
                        if event.end_date:
                            date_str += f" - {event.end_date.strftime('%H:%M' if event.date.date() == event.end_date.date() else '%Y.%m.%d %H:%M')}"

                    responsible_names = event.responsible or ""

                    worksheet.cell(row=current_row, column=1, value=date_str).alignment = left_alignment
                    worksheet.cell(row=current_row, column=2, value=event.name).alignment = left_alignment
                    worksheet.cell(row=current_row, column=3, value=event.place or "").alignment = left_alignment
                    worksheet.cell(row=current_row, column=4,
                                   value=event.department.name if event.department else "").alignment = left_alignment
                    worksheet.cell(row=current_row, column=5, value=responsible_names).alignment = left_alignment

                    for col_num in range(1, 6):
                        cell = worksheet.cell(row=current_row, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    current_row += 1

        for row in worksheet.iter_rows(min_row=header_row, max_row=current_row - 1, max_col=5):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    char_count = len(str(cell.value))
                    estimated_height = min(100, max(15, lines * 15 + (char_count // 100) * 5))
                    worksheet.row_dimensions[cell.row].height = estimated_height

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = 'attachment; filename="План мероприятий.xlsx"'
        return response

    export_to_excel.short_description = "Экспортировать в Excel"

    actions = [export_to_excel]