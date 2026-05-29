from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from .forms import EventForm
from django.shortcuts import redirect
from django.http import HttpResponse
from django.views.decorators.http import require_POST
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from collections import defaultdict
from django.shortcuts import render
from django.core.paginator import Paginator
from datetime import datetime as dt
from datetime import datetime
from django.http import JsonResponse
from django.db.models import Q
import json
from django.utils import timezone
from .models import Event, EventDate


def event_list(request):
    events = Event.objects.all()
    return render(request, 'events/event_list.html', {'events': events})


def create_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)

            # Устанавливаем user_id
            if request.user and hasattr(request.user, 'id'):
                event.user_id = request.user.id
            else:
                event.user = None

            # Сохраняем мероприятие, чтобы получить ID
            event.save()

            # Получаем JSON с датами
            dates_json = request.POST.get('dates_json', '[]')
            try:
                dates_data = json.loads(dates_json)
            except:
                dates_data = []

            # Удаляем старые даты (если были)
            event.event_dates.all().delete()

            # Создаём новые даты в таблице EventDate
            for d in dates_data:
                try:
                    start_dt = datetime.strptime(d['start'], '%Y-%m-%dT%H:%M')
                    end_dt = datetime.strptime(d['end'], '%Y-%m-%dT%H:%M') if d.get('end') else None

                    EventDate.objects.create(
                        event=event,
                        start=timezone.make_aware(start_dt) if start_dt else None,
                        end=timezone.make_aware(end_dt) if end_dt else None
                    )
                except Exception as e:
                    print(f"Ошибка создания даты: {e}")

            # Устанавливаем первую дату как основную (для обратной совместимости)
            if dates_data:
                first_date = dates_data[0]
                try:
                    event.date = timezone.make_aware(
                        datetime.strptime(first_date['start'], '%Y-%m-%dT%H:%M')
                    )
                    if first_date.get('end'):
                        event.end_date = timezone.make_aware(
                            datetime.strptime(first_date['end'], '%Y-%m-%dT%H:%M')
                        )
                    event.save()
                except:
                    pass

            # Очищаем комментарий от старых дат (оставляем только чистый комментарий)
            original_comment = form.cleaned_data.get('comment', '') or ''
            if "ДАТЫ:" in original_comment:
                original_comment = original_comment.split("\n\nДАТЫ:")[0].strip()
            event.comment = original_comment
            event.save()

            return redirect('events_ui')
    else:
        form = EventForm()
        event_dates_json = '[]'

    return render(request, 'events/create_event.html', {'form': form, 'event_dates_json': event_dates_json})


def delete_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    event.delete()
    return redirect('events_ui')


def edit_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    # Формируем JSON для множественных дат
    event_dates = []
    for ed in event.event_dates.all():
        event_dates.append({
            'start': ed.start.strftime('%Y-%m-%dT%H:%M'),
            'end': ed.end.strftime('%Y-%m-%dT%H:%M') if ed.end else ''
        })
    event_dates_json = json.dumps(event_dates)

    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            event = form.save(commit=False)

            # Получаем JSON с датами
            dates_json = request.POST.get('dates_json', '[]')
            try:
                dates_data = json.loads(dates_json)
            except:
                dates_data = []

            # Удаляем старые даты
            event.event_dates.all().delete()

            # Создаём новые даты в таблице EventDate
            for d in dates_data:
                try:
                    start_dt = datetime.strptime(d['start'], '%Y-%m-%dT%H:%M')
                    end_dt = datetime.strptime(d['end'], '%Y-%m-%dT%H:%M') if d.get('end') else None

                    EventDate.objects.create(
                        event=event,
                        start=timezone.make_aware(start_dt) if start_dt else None,
                        end=timezone.make_aware(end_dt) if end_dt else None
                    )
                except Exception as e:
                    print(f"Ошибка создания даты: {e}")

            # Устанавливаем первую дату как основную (для обратной совместимости)
            if dates_data:
                first_date = dates_data[0]
                try:
                    event.date = timezone.make_aware(
                        datetime.strptime(first_date['start'], '%Y-%m-%dT%H:%M')
                    )
                    if first_date.get('end'):
                        event.end_date = timezone.make_aware(
                            datetime.strptime(first_date['end'], '%Y-%m-%dT%H:%M')
                        )
                except:
                    pass

            # Очищаем комментарий от старых дат
            original_comment = form.cleaned_data.get('comment', '') or ''
            if "ДАТЫ:" in original_comment:
                original_comment = original_comment.split("\n\nДАТЫ:")[0].strip()
            event.comment = original_comment
            event.save()

            return redirect('events_ui')
    else:
        form = EventForm(instance=event)

    return render(request, 'events/edit_event.html', {'form': form, 'event': event, 'event_dates_json': event_dates_json})


def export_events_to_excel(request):
    from io import BytesIO
    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Мероприятия"

    headers = [
        "ID", "Название", "Дата начала", "Дата окончания", "Категория",
        "Подразделение", "Ответственный", "Место", "Создатель", "Комментарий"
    ]
    worksheet.append(headers)

    events = Event.objects.all()
    for event in events:
        worksheet.append([
            event.id,
            event.name or "",
            event.date.strftime('%Y-%m-%d %H:%M') if event.date else "",
            event.end_date.strftime('%Y-%m-%d %H:%M') if event.end_date else "",
            event.category.name if event.category else "",
            event.department.name if event.department else "",
            event.responsible or "",
            event.place or "",
            event.user.username if event.user else "",
            event.comment or "",
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename="events_export.xlsx"'
    return response


def bulk_delete_events(request):
    if request.method == 'POST':
        selected_events = request.POST.getlist('selected_events')
        if selected_events:
            try:
                event_ids = [int(id) for id in selected_events]
                Event.objects.filter(id__in=event_ids).delete()

                from django.contrib import messages
                count = len(selected_events)
                if count == 1:
                    messages.success(request, 'Мероприятие успешно удалено!')
                else:
                    messages.success(request, f'{count} мероприятий успешно удалено!')
            except (ValueError, Event.DoesNotExist):
                messages.error(request, 'Ошибка при удалении мероприятий')

        return redirect('events_ui')


def events_ui(request):
    current_year = dt.now().year

    month_str = request.GET.get('month', '')
    year_str = request.GET.get('year', '')
    sort_order = request.GET.get('sort_order', 'desc')
    category_id = request.GET.get('category', '')
    department_id = request.GET.get('department', '')
    search_query = request.GET.get('search', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    responsible_query = request.GET.get('responsible', '')

    events = Event.objects.all().select_related('category', 'department').prefetch_related('event_dates')

    if category_id and category_id != '':
        try:
            category_int = int(category_id)
            events = events.filter(category_id=category_int)
        except ValueError:
            pass

    if department_id and department_id != '':
        try:
            department_int = int(department_id)
            events = events.filter(department_id=department_int)
        except ValueError:
            pass

    if responsible_query:
        events = events.filter(responsible__icontains=responsible_query)

    if start_date_str or end_date_str:
        if start_date_str:
            try:
                start_date = dt.strptime(start_date_str, '%Y-%m-%d')
                start_datetime = timezone.make_aware(start_date)
                events = events.filter(
                    Q(date__gte=start_datetime) |
                    Q(end_date__gte=start_datetime) |
                    Q(date__lte=start_datetime, end_date__gte=start_datetime)
                )
            except ValueError:
                pass

        if end_date_str:
            try:
                end_date = dt.strptime(end_date_str, '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                end_datetime = timezone.make_aware(end_date)
                events = events.filter(
                    Q(date__lte=end_datetime) |
                    Q(end_date__lte=end_datetime) |
                    Q(date__lte=end_datetime, end_date__gte=end_datetime)
                )
            except ValueError:
                pass

    if search_query:
        search_terms = search_query.split()
        search_q = Q()
        for term in search_terms:
            search_q |= (
                Q(name__icontains=term) |
                Q(place__icontains=term) |
                Q(comment__icontains=term) |
                Q(responsible__icontains=term) |
                Q(category__name__icontains=term) |
                Q(department__name__icontains=term)
            )
        events = events.filter(search_q)

    if sort_order == 'asc':
        events = events.order_by('date', 'name')
    else:
        events = events.order_by('-date', 'name')

    total_events = Event.objects.count()

    paginator = Paginator(events, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    for event in page_obj:
        if event.responsible:
            event.responsible_list = [r.strip() for r in event.responsible.split(',') if r.strip()]
        else:
            event.responsible_list = []

    from categories.models import Category
    from departments.models import Department

    categories = Category.objects.all().order_by('name')
    departments = Department.objects.all().order_by('name')

    context = {
        'page_obj': page_obj,
        'sort_order': sort_order,
        'current_year': current_year,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_department': department_id,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'categories': categories,
        'departments': departments,
        'total_events': total_events,
        'filtered_count': events.count(),
    }

    return render(request, 'events/eventsUI.html', context)


@csrf_exempt
@require_POST
def export_selected_events(request):
    selected_events_ids = request.POST.getlist('selected_events')
    if not selected_events_ids:
        return HttpResponse("Не выбрано ни одного мероприятия", status=400)

    queryset = Event.objects.filter(id__in=selected_events_ids)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "План мероприятий"

    column_widths = {
        'A': 15,
        'C': 40,
        'D': 35,
        'E': 30,
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    for _ in range(3):
        current_row += 1

    worksheet.merge_cells(f'D{current_row}:E{current_row}')
    worksheet.cell(row=current_row, column=4, value='УТВЕРЖДАЮ').alignment = center_alignment
    worksheet.cell(row=current_row, column=4).font = Font(bold=True)
    current_row += 1

    worksheet.merge_cells(f'D{current_row}:E{current_row}')
    worksheet.cell(row=current_row, column=4, value='Ректор ВоГУ').alignment = center_alignment
    current_row += 1
    worksheet.merge_cells(f'D{current_row}:E{current_row}')
    worksheet.cell(row=current_row, column=4, value='').alignment = center_alignment
    current_row += 1

    worksheet.cell(row=current_row, column=4, value='').alignment = center_alignment
    worksheet.cell(row=current_row, column=5, value='Д. В. Дворников').alignment = center_alignment
    current_row += 1

    worksheet.cell(row=current_row, column=4, value='').alignment = center_alignment
    worksheet.cell(row=current_row, column=5, value='2025 года').alignment = center_alignment
    current_row += 2

    worksheet.merge_cells(f'A{current_row}:E{current_row}')
    worksheet.cell(row=current_row, column=1, value='ПЛАН МЕРОПРИЯТИЙ УНИВЕРСИТЕТА')
    worksheet.cell(row=current_row, column=1).alignment = center_alignment
    worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 2

    headers = [
        "Дата, время",
        "Наименование мероприятия",
        "Место проведения мероприятия",
        "Структурное подразделение",
        "Ответственный работник"
    ]

    header_row = current_row
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    current_row += 1

    events_by_category = defaultdict(list)

    for event in queryset.order_by('category__name', 'date'):
        if event.category:
            events_by_category[event.category.name].append(event)
        else:
            events_by_category['Без категории'].append(event)

    for category_name, events in events_by_category.items():
        if category_name and events:
            worksheet.merge_cells(f'A{current_row}:E{current_row}')
            category_cell = worksheet.cell(row=current_row, column=1, value=category_name.upper())
            category_cell.font = Font(bold=True, size=11)
            category_cell.alignment = left_alignment
            current_row += 1

            for event in events:
                date_str = ""
                if event.date:
                    date_str = event.date.strftime('%Y.%m.%d %H:%M')
                    if event.end_date:
                        date_str += f" - {event.end_date.strftime('%H:%M' if event.date.date() == event.end_date.date() else '%Y.%m.%d %H:%M')}"

                responsible_names = event.responsible or ""

                worksheet.cell(row=current_row, column=1, value=date_str).alignment = left_alignment
                worksheet.cell(row=current_row, column=2, value=event.name).alignment = left_alignment
                worksheet.cell(row=current_row, column=3, value=event.place or "").alignment = left_alignment
                worksheet.cell(row=current_row, column=4,
                               value=event.department.name if event.department else "").alignment = left_alignment
                worksheet.cell(row=current_row, column=5, value=responsible_names).alignment = left_alignment

                for col_num in range(1, 6):
                    cell = worksheet.cell(row=current_row, column=col_num)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                current_row += 1

    for row in worksheet.iter_rows(min_row=header_row, max_row=current_row - 1, max_col=5):
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                char_count = len(str(cell.value))
                estimated_height = min(100, max(15, lines * 15 + (char_count // 100) * 5))
                worksheet.row_dimensions[cell.row].height = estimated_height

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename="План мероприятий.xlsx"'
    return response


def check_database(request):
    results = {
        'total_events': Event.objects.count(),
        'events_with_date': Event.objects.filter(date__isnull=False).count(),
        'events_with_end_date': Event.objects.filter(end_date__isnull=False).count(),
        'month_distribution': {},
        'sample_events': []
    }

    for month in range(1, 13):
        count = Event.objects.filter(
            Q(date__month=month) | Q(end_date__month=month)
        ).count()
        results['month_distribution'][month] = count

    sample = Event.objects.filter(date__isnull=False).order_by('date')[:5]
    for event in sample:
        results['sample_events'].append({
            'id': event.id,
            'name': event.name,
            'date': event.date,
            'date_month': event.date.month if event.date else None,
            'end_date': event.end_date,
            'end_date_month': event.end_date.month if event.end_date else None,
        })

    return JsonResponse(results)


def get_filtered_event_ids(request):
    try:
        month_str = request.GET.get('month', '')
        year_str = request.GET.get('year', '')
        category_id = request.GET.get('category', '')
        department_id = request.GET.get('department', '')
        search_query = request.GET.get('search', '')
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')

        events = Event.objects.all()

        month_int = None
        year_int = None

        if month_str and month_str.strip() and month_str != 'None':
            try:
                month_int = int(month_str)
                if not (1 <= month_int <= 12):
                    month_int = None
            except (ValueError, TypeError):
                pass

        if year_str and year_str.strip() and year_str != 'None':
            try:
                year_int = int(year_str)
            except (ValueError, TypeError):
                year_int = None

        if month_int and year_int:
            events = events.filter(
                Q(date__month=month_int, date__year=year_int) |
                Q(end_date__month=month_int, end_date__year=year_int)
            )
        elif year_int:
            events = events.filter(
                Q(date__year=year_int) | Q(end_date__year=year_int)
            )
        elif month_int:
            current_year_for_filter = dt.now().year
            events = events.filter(
                Q(date__month=month_int, date__year=current_year_for_filter) |
                Q(end_date__month=month_int, end_date__year=current_year_for_filter)
            )

        if category_id and category_id != '':
            try:
                category_int = int(category_id)
                events = events.filter(category_id=category_int)
            except ValueError:
                pass

        if department_id and department_id != '':
            try:
                department_int = int(department_id)
                events = events.filter(department_id=department_int)
            except ValueError:
                pass

        if start_date_str or end_date_str:
            if start_date_str:
                try:
                    start_date = dt.strptime(start_date_str, '%Y-%m-%d')
                    start_datetime = timezone.make_aware(start_date)
                    events = events.filter(
                        Q(date__gte=start_datetime) |
                        Q(end_date__gte=start_datetime) |
                        Q(date__lte=start_datetime, end_date__gte=start_datetime)
                    )
                except ValueError:
                    pass

            if end_date_str:
                try:
                    end_date = dt.strptime(end_date_str, '%Y-%m-%d')
                    end_date = end_date.replace(hour=23, minute=59, second=59)
                    end_datetime = timezone.make_aware(end_date)
                    events = events.filter(
                        Q(date__lte=end_datetime) |
                        Q(end_date__lte=end_datetime) |
                        Q(date__lte=end_datetime, end_date__gte=end_datetime)
                    )
                except ValueError:
                    pass

        if search_query:
            search_terms = search_query.split()
            search_q = Q()
            for term in search_terms:
                search_q |= (
                    Q(name__icontains=term) |
                    Q(place__icontains=term) |
                    Q(comment__icontains=term) |
                    Q(responsible__icontains=term) |
                    Q(category__name__icontains=term) |
                    Q(department__name__icontains=term)
                )
            events = events.filter(search_q)

        event_ids = list(events.values_list('id', flat=True))

        return JsonResponse({
            'event_ids': event_ids,
            'count': len(event_ids),
            'success': True
        })

    except Exception as e:
        return JsonResponse({
            'event_ids': [],
            'count': 0,
            'success': False,
            'error': str(e)
        }, status=500)